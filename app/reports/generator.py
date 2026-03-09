"""
ReportGenerator — executes pre-defined analytical queries and renders
results into PDF (ReportLab) or Excel (openpyxl) files.
"""
import logging
import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ReportLab — pure Python PDF generation, works on Windows with no system libs
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from app.database import get_pool
from app.reports import queries as Q

# Set up logging
logger = logging.getLogger(__name__)

# Directory where generated report files are stored temporarily
REPORTS_DIR = os.environ.get("REPORTS_DIR", "/tmp/mis_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportGenerator:
    """Generates PDF and Excel reports from analytical SQL queries."""

    # Pre-defined report types and their human-readable titles
    SUPPORTED_TYPES = {
        "weekly_sales": "Weekly Sales Summary",
        "monthly_revenue": "Monthly Revenue by Channel",
        "top_books": "Top Books by Units Sold",
        "sales_by_category": "Sales by Age Group / Category",
        "seasonal_trend": "Seasonal Sales Trend",
        "inventory": "Inventory Status Report",
    }

    async def generate(
        self,
        report_type: str,
        from_date: str,
        to_date: str,
        format: str = "pdf",
        channel_id: Optional[int] = None,
        category_id: Optional[int] = None,
        generated_by: Optional[int] = None,
    ) -> dict:
        """
        Generate a report and save to disk.
        Returns a dict with report metadata including file_path.
        """
        # Validate report type and format
        if report_type not in self.SUPPORTED_TYPES:
            raise ValueError(f"Unknown report type: {report_type}")

        # Get DB connection pool and fetch data for the report
        pool = await get_pool()
        rows, columns = await self._fetch_data(
            pool, report_type, from_date, to_date, channel_id, category_id
        )

        # Render report to file
        title = self.SUPPORTED_TYPES[report_type]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_{timestamp}.{format}"
        file_path = os.path.join(REPORTS_DIR, filename)

        # Render to Excel or PDF based on requested format
        if format == "xlsx":
            # Excel rendering with openpyxl
            self._render_excel(rows, columns, title, from_date, to_date, file_path)
        else:
            # PDF rendering with ReportLab
            self._render_pdf(rows, columns, title, from_date, to_date, file_path)

        # Persist report record in DB
        report_id = await self._save_report_record(
            pool, report_type, format, from_date, to_date,
            channel_id, category_id, file_path, generated_by
        )

        # Log the report generation event
        logger.info("Report generated: %s → %s", report_type, file_path)
        # Return metadata for API response
        return {
            "report_id": report_id,
            "file_path": file_path,
            "filename": filename,
            "report_type": report_type,
            "format": format,
            "generated_at": datetime.now().isoformat(),
        }

    # ── Data fetching ─────────────────────────────────────────────────────────

    async def _fetch_data(
        self, pool, report_type, from_date, to_date, channel_id, category_id
    ) -> tuple[list[dict], list[str]]:
        '''Fetch data for the specified report type and parameters.'''
        # asyncpg requires date objects, not strings
        from datetime import date as date_type
        # Convert from_date and to_date from strings to date objects
        def _to_date(v):
            if isinstance(v, date_type):
                return v
            return date_type.fromisoformat(str(v)) if v else None

        # Convert input date strings to date objects for query parameters
        fd = _to_date(from_date)
        td = _to_date(to_date)

        # Map report types to their corresponding SQL queries and parameters
        query_map = {
            "weekly_sales":      (Q.WEEKLY_SALES_SUMMARY, {}),
            "monthly_revenue":   (Q.REVENUE_BY_CHANNEL,   {"from_date": fd, "to_date": td}),
            "top_books":         (Q.TOP_BOOKS,             {"from_date": fd, "to_date": td, "channel_id": channel_id, "limit": 50}),
            "sales_by_category": (Q.SALES_BY_CATEGORY,    {"from_date": fd, "to_date": td, "channel_id": channel_id}),
            "seasonal_trend":    (Q.SEASONAL_TREND,        {"from_date": fd, "to_date": td, "channel_id": channel_id}),
            "inventory":         (Q.INVENTORY_STATUS,      {"status": "active", "category_id": category_id}),
        }
        # Get the SQL and parameters for the requested report type
        sql, params = query_map[report_type]

        # Replace named :param placeholders with $N for asyncpg
        positional_sql, values = self._named_to_positional(sql, params)

        # Execute the query and fetch results
        async with pool.acquire() as conn:
            rows = await conn.fetch(positional_sql, *values)

        # If no rows returned, return empty data and columns
        if not rows:
            return [], []

        # Convert asyncpg Record objects to list of dicts and extract column names
        columns = list(rows[0].keys())
        data = [dict(r) for r in rows]
        # Return the data and column names for rendering
        return data, columns

    @staticmethod
    def _named_to_positional(sql: str, params: dict) -> tuple[str, list]:
        """Convert :name style params to $1..$N for asyncpg."""
        values = []
        i = 1
        for key, value in params.items():
            if f":{key}" in sql:
                sql = sql.replace(f":{key}", f"${i}")
                values.append(value)
                i += 1
        return sql, values

    # ── Excel rendering ───────────────────────────────────────────────────────

    def _render_excel(
        self, rows, columns, title, from_date, to_date, file_path
    ) -> None:
        '''Render the report data into an Excel file using openpyxl.'''
        # Create a new workbook and set up the main sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Populating header row
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        # Write column headers with styling
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Populating data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        # Auto-fit column widths (approximate)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # Metadata sheet
        meta_ws = wb.create_sheet("Report Info")
        meta_ws["A1"] = "Report"
        meta_ws["B1"] = title
        meta_ws["A2"] = "Period"
        meta_ws["B2"] = f"{from_date} to {to_date}"
        meta_ws["A3"] = "Generated"
        meta_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_ws["A4"] = "System"
        meta_ws["B4"] = "MyEnglishBooks MIS"

        # Save the workbook to the specified file path
        wb.save(file_path)

    # ── PDF rendering (ReportLab) ─────────────────────────────────────────────

    def _render_pdf(
        self, rows, columns, title, from_date, to_date, file_path
    ) -> None:
        '''Render the report data into a PDF file using ReportLab.'''
        # Use landscape for wide tables, portrait for narrow ones
        pagesize = landscape(A4) if len(columns) > 6 else A4
        # Set up the PDF document with margins and styles
        doc = SimpleDocTemplate(
            file_path,
            pagesize=pagesize,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        # Get default styles and prepare the story (content) list
        styles = getSampleStyleSheet()
        story = []

        # ── Title ─────────────────────────────────────────────────────────────
        title_style = styles["Title"]
        story.append(Paragraph(f"📚 {title}", title_style))
        story.append(Spacer(1, 0.3 * cm))

        # ── Metadata line ─────────────────────────────────────────────────────
        meta_style = styles["Normal"]
        meta_style.fontSize = 9
        meta_style.textColor = colors.grey
        story.append(Paragraph(
            f"Period: {from_date} — {to_date} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            meta_style,
        ))
        story.append(Spacer(1, 0.5 * cm))

        # ── Data table ───────────────────────────────────────────────────────────
        if not rows:
            # If no data, show a placeholder message instead of an empty table
            story.append(Paragraph("No data available for the selected period.", styles["Normal"]))
        else:
            # ── Table header + data rows ───────────────────────────────────────
            header = [col.replace("_", " ").title() for col in columns]
            table_data = [header] + [
                [str(row.get(col, "") or "") for col in columns]
                for row in rows
            ]

            # Calculate column widths to fill the page
            page_w = pagesize[0] - 3 * cm  # total usable width
            col_w = page_w / len(columns)

            # Create the table with styling for header and alternating row colors
            table = Table(table_data, colWidths=[col_w] * len(columns), repeatRows=1)
            table.setStyle(TableStyle([
                # Header row
                ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1F4E79")),
                ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, 0),  9),
                ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
                ("BOTTOMPADDING",(0, 0), (-1, 0),  6),
                # Data rows
                ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE",     (0, 1), (-1, -1), 8),
                ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FC")]),
                ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
                ("TOPPADDING",   (0, 1), (-1, -1), 4),
                ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
            ]))
            story.append(table)

        # ── Footer ────────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.8 * cm))
        footer_style = styles["Normal"]
        footer_style.fontSize = 8
        footer_style.textColor = colors.lightgrey
        story.append(Paragraph("MyEnglishBooks MIS — Confidential", footer_style))

        # Build the PDF document with the story content
        doc.build(story)

    # ── DB record ─────────────────────────────────────────────────────────────

    async def _save_report_record(
        self, pool, report_type, format, from_date, to_date,
        channel_id, category_id, file_path, generated_by
    ) -> int:
        '''Save a record of the generated report in the database and return its ID.'''
        import json
        # Store the report parameters as JSON in the database for auditing and future reference
        params_json = json.dumps({
            "from_date": from_date, "to_date": to_date,
            "channel_id": channel_id, "category_id": category_id,
        })
        # Insert a new record into the mis.scheduled_report table with the report metadata
        async with pool.acquire() as conn:
            # This SQL statement inserts a new report record and returns the generated report_id.
            # The generated_at timestamp is set to NOW() in the database, and the generated_by is the user ID of the requester
            row = await conn.fetchrow(
                """
                INSERT INTO mis.scheduled_report
                    (report_type, format, parameters, file_path,
                     generated_at, generated_by)
                VALUES ($1, $2, $3::jsonb, $4, NOW(), $5)
                RETURNING report_id
                """,
                report_type, format, params_json, file_path, generated_by,
            )
        return row["report_id"]

